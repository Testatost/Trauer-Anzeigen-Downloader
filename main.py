# -*- coding: utf-8 -*-
"""
Traueranzeigen-Downloader (Multi-Sites) – PySide6

Unterstützte Domains:
- https://trauer-anzeigen.de/
- https://www.abschied-nehmen.de/   (+ abschied-nehmen.de)
- https://www.ok-trauer.de/         (+ ok-trauer.de)
- https://gedenken.freiepresse.de/
- https://www.vrm-trauer.de/        (+ vrm-trauer.de)

Input oben:
- URL (http/https) -> normaler Crawl ab dieser URL
- sonst -> Name-Modus: sucht auf allen unterstützten Seiten über deren Such-URL:
  /traueranzeigen-suche/<slug>

Features:
- Modus: Bilder / Daten / Bilder+Daten
- Export: XLSX / JSON / CSV / XML (URL immer letzte Spalte/Key)
- Suche/Filter im Table (lokal)
- Max. Personen: Input -> Crawler stoppt automatisch
- OCR manuell (markierte Zeilen), schreibt alles in "zusatzinformationen"
- Shortcuts: Entf / Ctrl+S / Ctrl+Q / Ctrl+A / Ctrl+O
- Buttons: Auswahl löschen / Reset / Speichern
- Tesseract fehlt -> Hilfe-Dialog mit klickbarem TU Mannheim Link
- Tabellen-Sortierung per Klick auf Spaltenkopf (A-Z / Z-A)

Wichtige Änderungen (Fixes):
1) Name-Suche: Parsing ist jetzt wieder "result-orientiert":
   - Primär werden Ergebnis-Überschriften ("Anzeige <Name>") aus h1-h6 gelesen (wie bei den Portalen üblich).
   - Fallback greift nur auf sehr wahrscheinliche Detail-Links zurück.
   => Dadurch verschwinden "Livechat / Vorsorge / Aufgeben" praktisch komplett.

2) Kein Duplikat-Filter (wie gewünscht), weil es mehrere Anzeigen pro Person geben kann.

3) Sortierung: nutzt jetzt Qt.EditRole als Sort-Key (damit PLZ / Datum sauber sortieren).
"""

from __future__ import annotations

import csv
import json
import os
import random
import re
import time
import traceback
from dataclasses import dataclass
from io import BytesIO
from typing import Optional, List, Tuple
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from PIL import Image, ImageOps, ImageFilter

# optional anti-bot
try:
    import cloudscraper
except Exception:
    cloudscraper = None

# optional OCR
try:
    import pytesseract
except Exception:
    pytesseract = None

import openpyxl
from openpyxl.utils import get_column_letter

from PySide6 import QtCore, QtWidgets
from PySide6.QtGui import QDesktopServices, QKeySequence, QShortcut
from PySide6.QtCore import QUrl

import shutil


# ---------------- Config ----------------

USER_AGENT_LIST = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129 Safari/605.1",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
]

ALLOWED_HOSTS = {
    "trauer-anzeigen.de",
    "www.abschied-nehmen.de",
    "abschied-nehmen.de",
    "www.ok-trauer.de",
    "ok-trauer.de",
    "gedenken.freiepresse.de",
    "www.vrm-trauer.de",
    "vrm-trauer.de",
}

SEARCH_HOSTS_ORDERED = [
    "trauer-anzeigen.de",
    "abschied-nehmen.de",
    "www.abschied-nehmen.de",
    "ok-trauer.de",
    "www.ok-trauer.de",
    "gedenken.freiepresse.de",
    "vrm-trauer.de",
    "www.vrm-trauer.de",
]

TESSERACT_CANDIDATES = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
]

TESSERACT_DOWNLOAD_URL = "https://digi.bib.uni-mannheim.de/tesseract/tesseract-ocr-w64-setup-5.4.0.20240606.exe"

# Wörter/Strings, die KEINE echten Personentreffer sind (Buttons/Navi/Marketing)
JUNK_TITLES = {
    "suchen",
    "aufgeben",
    "online aufgeben",
    "anzeige aufgeben",
    "trau­er­an­zei­ge auf­ge­ben",
    "jetzt neu",
    "video",
    "video-anzeige",
    "die video-anzeige",
    "traueranzeige aufgeben",
    "augeben",  # Tippfehler-Fälle
    "aufgeben!",
    "anzeigen aufgeben",
    "mehr",
    "zum livechat »",
    "trauerhilfe live-chat",
    "cookie-einstellungen",
    "schließen",
}

# Einige Portale zeigen diese CTAs/Teaser; wir filtern zusätzlich per Muster
JUNK_PATTERNS = [
    r"\baufgeben\b",
    r"\banzeige\s+aufgeben\b",
    r"\bsuchen\b",
    r"\bvideo\b",
    r"\bjetzt\b.*\bneu\b",
    r"\blivechat\b",
    r"\btrauerchat\b",
    r"\bvorsorge\b",
    r"\btestament\b",
    r"\berbe\b",
    r"\bkondolieren\b",
    r"\btrauerhilfe\b",
    r"\bim todesfall\b",
    r"\bdatenschutz\b",
    r"\bimpressum\b",
    r"\bagb\b",
]


def configure_tesseract(log_cb=print) -> bool:
    if not pytesseract:
        log_cb("OCR: pytesseract nicht installiert.")
        return False

    p = shutil.which("tesseract")
    if p and os.path.isfile(p):
        pytesseract.pytesseract.tesseract_cmd = p
        log_cb(f"OCR: tesseract gefunden im PATH: {p}")
        return True

    for c in TESSERACT_CANDIDATES:
        if os.path.isfile(c):
            pytesseract.pytesseract.tesseract_cmd = c
            log_cb(f"OCR: tesseract gefunden: {c}")
            return True

    log_cb("OCR: tesseract.exe nicht gefunden. Pfad manuell setzen (TESSERACT_CANDIDATES).")
    return False


def safe_filename(name: str) -> str:
    return re.sub(r"[^\w\-.]+", "_", (name or ""))[:180]


def split_name(full: str) -> tuple[str, str]:
    full = (full or "").strip()
    parts = full.split()
    if len(parts) >= 2:
        return " ".join(parts[:-1]), parts[-1]
    return full, ""


def normalize_date_de(s: str) -> str:
    s = (s or "").strip()
    m = re.search(r"([0-3]?\d\.[01]?\d\.\d{4})", s)
    return m.group(1) if m else ""


def normalize_name_key(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\wäöüß \-]", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def slugify_search_term(term: str) -> str:
    """
    'Max Mustermann' -> 'max-mustermann'
    Umlaute: ä->ae ö->oe ü->ue ß->ss (sehr häufig bei Portalen)
    """
    s = (term or "").strip().lower()
    s = s.replace("ß", "ss")
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    s = re.sub(r"[^a-z0-9\s\-]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s


def build_search_urls_for_host(host: str, name_query: str) -> List[str]:
    slug = slugify_search_term(name_query)
    if not slug:
        return []
    u1 = f"https://{host}/traueranzeigen-suche/{slug}"
    u2 = u1 + "/"
    out = []
    for u in (u1, u2):
        if u not in out:
            out.append(u)
    return out


def make_session() -> requests.Session:
    if cloudscraper:
        try:
            s = cloudscraper.create_scraper(browser={"browser": "chrome", "platform": "windows"})
            s.headers.update({"User-Agent": random.choice(USER_AGENT_LIST)})
            return s
        except Exception:
            pass
    s = requests.Session()
    s.headers.update({"User-Agent": random.choice(USER_AGENT_LIST)})
    return s


# ---------------- Data ----------------

@dataclass
class EntryRecord:
    name: str = ""
    nachname: str = ""
    full_name: str = ""
    plz: str = ""
    ort: str = ""
    geburtsdatum: str = ""
    sterbedatum: str = ""

    bild_url: str = ""
    bild_datei: str = ""  # intern nötig fürs OCR/Download, wird NICHT exportiert

    zusatzinformationen: str = ""  # alles OCR-te rein

    detail_url: str = ""  # einzige URL (immer letzte Spalte/Key)


class CrawlOptions(QtCore.QObject):
    def __init__(self, mode: str, download_dir: str, max_pages: int = 50, max_people: int = 0):
        super().__init__()
        self.mode = mode  # "images" | "data" | "both"
        self.download_dir = download_dir
        self.max_pages = max_pages
        self.max_people = max_people  # 0 = unendlich


# ---------------- OCR ----------------

def _to_grayscale(im: Image.Image) -> Image.Image:
    if im.mode in ("RGBA", "P"):
        im = im.convert("RGB")
    if im.mode != "L":
        im = im.convert("L")
    return im


def _adaptive_threshold(im_l: Image.Image) -> Image.Image:
    im = im_l.copy()
    im = im.filter(ImageFilter.MedianFilter(size=3))
    im = ImageOps.autocontrast(im)

    hist = im.histogram()
    total = sum(hist)
    if total == 0:
        return im

    cum = 0
    thresh = 160
    for i, c in enumerate(hist):
        cum += c
        if cum / total >= 0.55:
            thresh = i
            break

    thresh = max(120, min(200, thresh))
    im = im.point(lambda p: 255 if p > thresh else 0)
    return im


def preprocess_variants(im: Image.Image) -> List[Image.Image]:
    im0 = im
    if im0.mode in ("RGBA", "P"):
        im0 = im0.convert("RGB")

    w, h = im0.size
    scale = 3 if max(w, h) < 1600 else 2
    im_up = im0.resize((w * scale, h * scale), Image.Resampling.LANCZOS)

    variants: List[Image.Image] = []
    g = _to_grayscale(im_up)

    v1 = ImageOps.autocontrast(g)
    v1 = v1.filter(ImageFilter.UnsharpMask(radius=2, percent=220, threshold=3))
    variants.append(v1)

    v2 = _adaptive_threshold(g)
    variants.append(v2)

    v3 = g.filter(ImageFilter.MedianFilter(size=5))
    v3 = ImageOps.autocontrast(v3)
    v3 = v3.filter(ImageFilter.UnsharpMask(radius=3, percent=250, threshold=2))
    v3 = _adaptive_threshold(v3)
    variants.append(v3)

    return variants


def score_ocr_text(text: str) -> float:
    t = (text or "").strip()
    if not t:
        return 0.0

    letters = len(re.findall(r"[A-Za-zÄÖÜäöüß]", t))
    digits = len(re.findall(r"\d", t))
    words = len(re.findall(r"\b\w+\b", t))
    dates = len(re.findall(r"\b\d{1,2}\.\d{1,2}\.\d{4}\b", t))
    common = len(re.findall(r"\b(Eltern|Kinder|Enkel|Geschwister|In\s+Liebe|Trauer|Danke|Beisetzung)\b", t, flags=re.I))
    garbage = len(re.findall(r"[#@{}\[\]\\^~_<>|]", t))

    score = 0.0
    score += letters * 0.8
    score += words * 1.2
    score += digits * 0.2
    score += dates * 25.0
    score += common * 15.0
    score -= garbage * 5.0

    if len(t) >= 120:
        score += 40.0
    elif len(t) >= 60:
        score += 20.0

    return score


def normalize_ocr_text(text: str) -> str:
    lines = []
    for ln in (text or "").splitlines():
        ln = ln.strip()
        if not ln:
            continue
        ln = re.sub(r"\s+", " ", ln)
        if len(ln) < 8 and len(re.findall(r"[A-Za-zÄÖÜäöüß]", ln)) < 3:
            continue
        lines.append(ln)
    return "\n".join(lines).strip()


def ocr_best_effort(img_bytes: bytes) -> str:
    if not pytesseract:
        return ""

    im0 = Image.open(BytesIO(img_bytes))
    if im0.mode in ("RGBA", "P"):
        im0 = im0.convert("RGB")

    w, h = im0.size
    rois: List[Tuple[str, Image.Image]] = [
        ("full", im0),
        ("bottom", im0.crop((0, int(h * 0.55), w, h))),
        ("middle", im0.crop((0, int(h * 0.25), w, int(h * 0.80)))),
    ]

    psms = [6, 4, 11]
    best_text = ""
    best_score = 0.0

    try:
        _ = pytesseract.image_to_osd(im0)
    except Exception:
        pass

    base_config = r'--oem 3 -c preserve_interword_spaces=1'

    for _, roi in rois:
        variants = preprocess_variants(roi)
        for vimg in variants:
            for psm in psms:
                cfg = f"{base_config} --psm {psm}"
                try:
                    txt = pytesseract.image_to_string(vimg, lang="deu", config=cfg) or ""
                except Exception:
                    continue
                txt_n = normalize_ocr_text(txt)
                sc = score_ocr_text(txt_n)
                if sc > best_score:
                    best_score = sc
                    best_text = txt_n

    return best_text.strip()


# ---------------- Crawler ----------------

class Crawler:
    def __init__(self, sess: requests.Session, opts: CrawlOptions, log_cb):
        self.sess = sess
        self.opts = opts
        self.log = log_cb
        self.seen_pages: set[str] = set()

    def _get(self, url: str, timeout: int = 25) -> tuple[str, str, int]:
        try:
            r = self.sess.get(url, timeout=timeout, allow_redirects=True)
            return (r.text or ""), (str(r.url) if r.url else url), int(r.status_code)
        except Exception:
            return "", url, 0

    def _validate_url(self, url: str) -> None:
        p = urlparse(url)
        if p.scheme not in ("http", "https"):
            raise ValueError("URL muss mit http/https beginnen.")
        host = (p.netloc or "").lower()
        if host not in ALLOWED_HOSTS:
            raise ValueError(
                "Nicht unterstützte Domain.\n"
                "Erlaubt: trauer-anzeigen.de, abschied-nehmen.de, ok-trauer.de, gedenken.freiepresse.de, vrm-trauer.de"
            )

    def _fetch_detail_soup(self, url: str) -> Optional[BeautifulSoup]:
        html, _, st = self._get(url, timeout=25)
        if st != 200 or not html:
            return None
        return BeautifulSoup(html, "html.parser")

    # --- Filters (gegen "aufgeben"/"suchen"/Teaser)

    def _is_junk_title(self, s: str) -> bool:
        t = (s or "").strip().lower()
        t = re.sub(r"\s+", " ", t)
        if not t:
            return True
        if t in JUNK_TITLES:
            return True
        for pat in JUNK_PATTERNS:
            if re.search(pat, t, flags=re.I):
                return True
        if len(t) <= 3:
            return True
        return False

    def _looks_like_detail_url(self, url: str) -> bool:
        """
        Heuristik: echte Anzeigen-Detailseiten haben meist:
        - /traueranzeige/ oder /trauerfall/ oder /nachruf/ oder /gedenkseite/
        - oder enthalten Zahlen/IDs
        """
        u = (url or "").lower()
        if not u.startswith("http"):
            return False

        # block obvious non-details
        if any(x in u for x in ("/kontakt", "/impressum", "/datenschutz", "/agb", "/login", "/register")):
            return False

        # block "create ad"/marketing
        if any(x in u for x in ("/aufgeben", "anzeige-aufgeben", "traueranzeige-aufgeben", "anzeigen-aufgeben")):
            return False

        # our search listing itself is not a detail page
        if "/traueranzeigen-suche/" in u:
            return False

        # allow typical patterns
        if "/traueranzeige" in u:
            return True
        if "/trauerfall" in u:
            return True
        if "/nachruf" in u:
            return True
        if "/gedenk" in u:  # gedenkseite / gedenken / etc.
            # nicht zu breit: dennoch ok als detail, weil Portale so benennen
            return True

        # contains an id-ish number
        if re.search(r"/\d{5,}(\b|/|$)", u):
            return True

        return False

    # --- image heuristics

    def _is_logoish(self, url: str) -> bool:
        u = (url or "").lower()
        bad = [
            "logo", "banner", "placeholder", "favicon", "sprite",
            "cookie", "consent", "header", "footer", "tracking", "analytics",
        ]
        return any(x in u for x in bad)

    def _looks_like_obituary_image(self, url: str) -> bool:
        u = (url or "").lower()
        if "traueranzeige" in u:
            return True
        if "anzeige" in u and re.search(r"\.(jpg|jpeg|png)(\?|$)", u):
            return True
        if re.search(r"(portrait|profil|profile)\.(jpg|jpeg|png)(\?|$)", u):
            return True
        return False

    def _extract_image_from_listing_container(self, container, base_url: str) -> str:
        best_fallback = ""
        for img in container.find_all("img"):
            cand = ""
            for attr in ("data-src", "data-original", "src", "srcset", "data-srcset"):
                v = img.get(attr)
                if not v:
                    continue
                cand = v.split()[0]
                break
            if not cand:
                continue

            abs_url = urljoin(base_url, cand)
            if self._is_logoish(abs_url):
                continue
            if self._looks_like_obituary_image(abs_url):
                return abs_url
            if not best_fallback:
                best_fallback = abs_url
        return best_fallback

    def _extract_image_from_detail(self, soup: BeautifulSoup, detail_url: str) -> str:
        for meta in (
            soup.find("meta", attrs={"property": "og:image"}),
            soup.find("meta", attrs={"name": "twitter:image"}),
        ):
            if meta and meta.get("content"):
                u = urljoin(detail_url, meta["content"])
                if not self._is_logoish(u):
                    return u

        link = soup.find("link", rel=lambda v: v and "image_src" in v)
        if link and link.get("href"):
            u = urljoin(detail_url, link["href"])
            if not self._is_logoish(u):
                return u

        best_fallback = ""
        for img in soup.find_all("img", src=True):
            u = urljoin(detail_url, img["src"])
            if self._is_logoish(u):
                continue
            if self._looks_like_obituary_image(u):
                return u
            if not best_fallback:
                best_fallback = u
        return best_fallback

    def _download_image_if_any(self, rec: EntryRecord) -> None:
        if not rec.bild_url:
            return
        if self._is_logoish(rec.bild_url):
            return
        try:
            r = self.sess.get(rec.bild_url, timeout=20)
            if r.status_code != 200 or not r.content:
                return

            im = Image.open(BytesIO(r.content))
            if im.mode in ("P",):
                im = im.convert("RGBA")
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGBA")

            folder = os.path.join(self.opts.download_dir, "bilder")
            os.makedirs(folder, exist_ok=True)

            fn = f"{safe_filename(rec.nachname)}_{safe_filename(rec.name)}_{safe_filename(rec.geburtsdatum)}_{safe_filename(rec.sterbedatum)}.png"
            fp = os.path.join(folder, fn)
            im.save(fp, format="PNG", optimize=True)
            rec.bild_datei = fp
            self.log(f"Bild gespeichert: {fp}")
        except Exception as ex:
            self.log(f"Bildfehler: {ex}")

    # --- detail enrichment (data)

    def _apply_jsonld(self, rec: EntryRecord, data) -> None:
        objs = data if isinstance(data, list) else [data]
        for obj in objs:
            if not isinstance(obj, dict):
                continue
            addr = obj.get("address")
            if isinstance(addr, dict):
                pc = addr.get("postalCode") or addr.get("postal_code")
                loc = addr.get("addressLocality") or addr.get("locality")
                if pc and not rec.plz:
                    rec.plz = str(pc).strip()
                if loc and not rec.ort:
                    rec.ort = str(loc).strip()

    def _enrich_from_detail_soup(self, rec: EntryRecord, soup: BeautifulSoup) -> None:
        for script in soup.find_all("script", type=re.compile(r"ld\+json", re.I)):
            try:
                data = json.loads(script.get_text(strip=True) or "{}")
                self._apply_jsonld(rec, data)
            except Exception:
                continue

        text = soup.get_text(" ", strip=True)
        m = re.search(r"\b(\d{5})\s+([A-ZÄÖÜ][A-Za-zÄÖÜäöüß\- ]{2,})\b", text)
        if m:
            if not rec.plz:
                rec.plz = m.group(1)
            if not rec.ort:
                rec.ort = m.group(2).strip()

    def _enrich_from_detail(self, rec: EntryRecord) -> None:
        soup = self._fetch_detail_soup(rec.detail_url)
        if not soup:
            return
        self._enrich_from_detail_soup(rec, soup)
        if not rec.bild_url:
            rec.bild_url = self._extract_image_from_detail(soup, rec.detail_url)

    # ---------- Pagination: generic "next link" ----------

    def _find_next_page_url(self, soup: BeautifulSoup, current_url: str) -> str:
        link = soup.find("link", rel=lambda v: v and "next" in v)
        if link and link.get("href"):
            return urljoin(current_url, link["href"])

        a = soup.find("a", rel=lambda v: v and "next" in v, href=True)
        if a:
            return urljoin(current_url, a["href"])

        next_texts = ["nächste", "naechste", "weiter", "»", "›", "next"]
        for a in soup.find_all("a", href=True):
            t = (a.get_text(" ", strip=True) or "").lower()
            if any(x in t for x in next_texts):
                return urljoin(current_url, a["href"])

        for a in soup.select("a.next, a.pagination-next, li.next a, a[aria-label*=Next], a[aria-label*=näch]"):
            if a.get("href"):
                return urljoin(current_url, a["href"])

        return ""

    def _iter_listing_pages(self, start_url: str, stop_flag: callable) -> List[tuple[str, str]]:
        pages: List[tuple[str, str]] = []
        url = start_url
        for _ in range(self.opts.max_pages):
            if stop_flag():
                break
            if not url or url in self.seen_pages:
                break
            html, final_url, st = self._get(url)
            if st != 200 or not html:
                break
            if final_url in self.seen_pages:
                break
            self.seen_pages.add(final_url)
            pages.append((html, final_url))
            soup = BeautifulSoup(html, "html.parser")
            nxt = self._find_next_page_url(soup, final_url)
            if not nxt or nxt in self.seen_pages:
                break
            url = nxt
            time.sleep(0.35)
        return pages

    # ---------- Normal Crawl by URL ----------

    def crawl(self, start_url: str, stop_flag: callable) -> list[EntryRecord]:
        self._validate_url(start_url)
        os.makedirs(self.opts.download_dir, exist_ok=True)

        results: list[EntryRecord] = []

        self.log(f"[{self.opts.mode}] Listing starten: {start_url}")
        self.seen_pages.clear()
        pages = self._iter_listing_pages(start_url, stop_flag=stop_flag)
        if not pages:
            return results

        for html, base_url in pages:
            if stop_flag():
                break
            if self.opts.max_people and len(results) >= self.opts.max_people:
                self.log(f"Stop: Max. Personen erreicht ({self.opts.max_people}).")
                break
            self._parse_listing_page(html, base_url, results, stop_flag)

        return results

    # ---------- Name Search across all sites (Suchseite nutzen) ----------

    def _matches_name_query(self, full_name: str, query: str) -> bool:
        q = normalize_name_key(query)
        namek = normalize_name_key(full_name)
        if not q or not namek:
            return False

        tokens = [t for t in q.split(" ") if t]
        # match: alle token müssen im Ergebnisnamen enthalten sein
        return all(t in namek for t in tokens)

    def crawl_all_sites_by_name(self, name_query: str, stop_flag: callable) -> list[EntryRecord]:
        if not (name_query or "").strip():
            raise ValueError("Bitte einen Namen eingeben (für Name-Modus).")

        os.makedirs(self.opts.download_dir, exist_ok=True)
        results: list[EntryRecord] = []

        self.log(f"[name-search] Suche nach: {name_query}")

        # Name-Suche soll nicht ewig paginieren:
        # In der Praxis liefern die Suchseiten die relevantesten Treffer auf den ersten Seiten.
        old_max_pages = self.opts.max_pages
        self.opts.max_pages = min(self.opts.max_pages, 12)

        try:
            for host in SEARCH_HOSTS_ORDERED:
                if stop_flag():
                    break
                if host not in ALLOWED_HOSTS:
                    continue

                search_urls = build_search_urls_for_host(host, name_query)
                if not search_urls:
                    continue

                self.log(f"[name-search] Domain: {host}")
                found_any = False

                for start_url in search_urls:
                    if stop_flag():
                        break
                    if self.opts.max_people and len(results) >= self.opts.max_people:
                        self.log(f"Stop: Max. Personen erreicht ({self.opts.max_people}).")
                        break

                    self.log(f"[name-search] Suche-URL: {start_url}")

                    self.seen_pages.clear()
                    pages = self._iter_listing_pages(start_url, stop_flag=stop_flag)
                    if not pages:
                        continue

                    for html, base_url in pages:
                        if stop_flag():
                            break
                        if self.opts.max_people and len(results) >= self.opts.max_people:
                            break

                        page_records: list[EntryRecord] = []
                        self._parse_listing_page(html, base_url, page_records, stop_flag)

                        # nur echte Matches sammeln
                        for rec in page_records:
                            if stop_flag():
                                break
                            if self.opts.max_people and len(results) >= self.opts.max_people:
                                break
                            if not self._matches_name_query(rec.full_name, name_query):
                                continue

                            if self.opts.mode in ("data", "both"):
                                try:
                                    self._enrich_from_detail(rec)
                                except Exception:
                                    pass

                            results.append(rec)
                            found_any = True
                            display = rec.full_name or (rec.name + " " + rec.nachname).strip() or rec.detail_url
                            self.log(f"[name-search] Match #{len(results)} ({host}): {display}")

                    time.sleep(0.25)

                if not found_any:
                    self.log(f"[name-search] Keine Treffer bei {host}.")

            self.log(f"[name-search] Fertig. Matches: {len(results)}")
            return results
        finally:
            self.opts.max_pages = old_max_pages

    # ---------- Listing Parse (Fix: Ergebnis-orientiert statt "alle Links") ----------

    def _parse_listing_page(self, html: str, base_url: str, results: list[EntryRecord], stop_flag: callable) -> int:
        soup = BeautifulSoup(html, "html.parser")
        new_entries = 0

        candidates: List[tuple[str, str, Optional[object]]] = []  # (detail_url, full_name, container_hint)

        # 1) Primär: Überschriften wie "Anzeige <Name>" (sehr zuverlässig bei den Portalen)
        for h in soup.find_all(re.compile(r"^h[1-6]$")):
            if stop_flag():
                break
            a = h.find("a", href=True)
            if not a:
                continue
            txt = (h.get_text(" ", strip=True) or "").strip()
            if not txt:
                continue
            if not re.search(r"\bAnzeige\b", txt, flags=re.I):
                continue

            full_name = re.sub(r"^\s*Anzeige\s*", "", txt, flags=re.I).strip()
            if self._is_junk_title(full_name):
                continue

            detail_url = urljoin(base_url, (a.get("href") or "").strip())
            if not self._looks_like_detail_url(detail_url):
                continue

            container = h.find_parent(["article", "li", "div"]) or h
            candidates.append((detail_url, full_name, container))

        # 2) Fallback: falls ein Portal keine "Anzeige"-Überschriften hat
        #    => NUR sehr wahrscheinliche Detail-Links, und Name aus title/aria/h-Text
        if not candidates:
            for a in soup.find_all("a", href=True):
                if stop_flag():
                    break
                abs_url = urljoin(base_url, (a.get("href") or "").strip())
                if not self._looks_like_detail_url(abs_url):
                    continue

                txt = (a.get_text(" ", strip=True) or "").strip()
                if not txt:
                    txt = (a.get("title") or a.get("aria-label") or "").strip()

                # harte Filter: keine CTA-Buttons etc.
                if self._is_junk_title(txt):
                    continue

                # wenn irgendwo "Anzeige" steht, bevorzugt als Name extrahieren
                if re.search(r"\bAnzeige\b", txt, flags=re.I):
                    txt = re.sub(r"^\s*Anzeige\s*", "", txt, flags=re.I).strip()

                # heuristisch: "Zum Trauerfall" ist kein Name
                if re.search(r"\bzum trauerfall\b", txt, flags=re.I):
                    continue

                if self._is_junk_title(txt):
                    continue

                container = a.find_parent(["article", "li", "div"]) or a
                candidates.append((abs_url, txt, container))

        # unique by URL (pro Seite reicht ein Eintrag je Detail-URL)
        uniq: List[tuple[str, str, Optional[object]]] = []
        seen_u = set()
        for u, t, c in candidates:
            if u in seen_u:
                continue
            seen_u.add(u)
            uniq.append((u, t, c))
        candidates = uniq

        # records bauen
        for detail_url, full_name, container in candidates:
            if stop_flag():
                break
            if self.opts.max_people and len(results) >= self.opts.max_people:
                break

            rec = EntryRecord(full_name=full_name, detail_url=detail_url)
            rec.name, rec.nachname = split_name(full_name)

            # progress log (für alle Modi)
            display_name = rec.full_name or (rec.name + " " + rec.nachname).strip() or rec.detail_url
            self.log(f"[{self.opts.mode}] #{len(results) + 1}: {display_name}")

            # Datum evtl. im Container
            try:
                block = container.get_text(" ", strip=True) if container else ""
            except Exception:
                block = ""

            m = re.search(r"(\d{1,2}\.\d{1,2}\.\d{4}).*?[–-].*?(\d{1,2}\.\d{1,2}\.\d{4})", block)
            if m:
                rec.geburtsdatum = normalize_date_de(m.group(1))
                rec.sterbedatum = normalize_date_de(m.group(2))

            # Bild aus Container, sonst ggf. später aus Detail
            if container is not None:
                rec.bild_url = self._extract_image_from_listing_container(container, base_url)

            if self.opts.mode == "images":
                if not rec.bild_url:
                    ds = self._fetch_detail_soup(rec.detail_url)
                    if ds:
                        rec.bild_url = self._extract_image_from_detail(ds, rec.detail_url)
                self._download_image_if_any(rec)

            elif self.opts.mode == "data":
                self._enrich_from_detail(rec)

            else:  # both
                ds = self._fetch_detail_soup(rec.detail_url)
                if ds:
                    if not rec.bild_url:
                        rec.bild_url = self._extract_image_from_detail(ds, rec.detail_url)
                    self._download_image_if_any(rec)
                    self._enrich_from_detail_soup(rec, ds)
                else:
                    self._download_image_if_any(rec)

            results.append(rec)
            new_entries += 1

        return new_entries


# ---------------- Exporter ----------------

EXPORT_HEADERS = [
    "name", "nachname", "plz", "ort", "geburtsdatum", "sterbedatum",
    "zusatzinformationen",
    "anzeige_url",
]


def export_xlsx(path: str, items: list[EntryRecord]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traueranzeigen"
    ws.append(EXPORT_HEADERS)
    for r in items:
        ws.append([
            r.name, r.nachname, r.plz, r.ort, r.geburtsdatum, r.sterbedatum,
            r.zusatzinformationen,
            r.detail_url,
        ])
    for i, _ in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 32
    wb.save(path)


def export_json(path: str, items: list[EntryRecord]) -> None:
    data = []
    for r in items:
        data.append({
            "name": r.name,
            "nachname": r.nachname,
            "plz": r.plz,
            "ort": r.ort,
            "geburtsdatum": r.geburtsdatum,
            "sterbedatum": r.sterbedatum,
            "zusatzinformationen": r.zusatzinformationen,
            "anzeige_url": r.detail_url,
        })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def export_csv(path: str, items: list[EntryRecord]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=EXPORT_HEADERS)
        w.writeheader()
        for r in items:
            w.writerow({
                "name": r.name,
                "nachname": r.nachname,
                "plz": r.plz,
                "ort": r.ort,
                "geburtsdatum": r.geburtsdatum,
                "sterbedatum": r.sterbedatum,
                "zusatzinformationen": r.zusatzinformationen,
                "anzeige_url": r.detail_url,
            })


def export_xml(path: str, items: list[EntryRecord]) -> None:
    def esc(s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&apos;")

    lines = ['<?xml version="1.0" encoding="UTF-8"?>', "<traueranzeigen>"]
    for r in items:
        lines.append("  <eintrag>")
        lines.append(f"    <name>{esc(r.name)}</name>")
        lines.append(f"    <nachname>{esc(r.nachname)}</nachname>")
        lines.append(f"    <plz>{esc(r.plz)}</plz>")
        lines.append(f"    <ort>{esc(r.ort)}</ort>")
        lines.append(f"    <geburtsdatum>{esc(r.geburtsdatum)}</geburtsdatum>")
        lines.append(f"    <sterbedatum>{esc(r.sterbedatum)}</sterbedatum>")
        lines.append(f"    <zusatzinformationen>{esc(r.zusatzinformationen)}</zusatzinformationen>")
        lines.append(f"    <anzeige_url>{esc(r.detail_url)}</anzeige_url>")
        lines.append("  </eintrag>")
    lines.append("</traueranzeigen>")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ---------------- Workers ----------------

class CrawlWorker(QtCore.QObject):
    log = QtCore.Signal(str)
    done = QtCore.Signal(list)
    failed = QtCore.Signal(str)

    def __init__(self, input_text: str, opts: CrawlOptions):
        super().__init__()
        self.input_text = input_text
        self.opts = opts
        self._stop = False

    def stop(self):
        self._stop = True

    @QtCore.Slot()
    def run(self):
        try:
            sess = make_session()
            crawler = Crawler(sess, self.opts, self.log.emit)

            txt = (self.input_text or "").strip()
            if re.match(r"^https?://", txt, flags=re.I):
                items = crawler.crawl(txt, stop_flag=lambda: self._stop)
            else:
                items = crawler.crawl_all_sites_by_name(txt, stop_flag=lambda: self._stop)

            self.done.emit(items)
        except Exception:
            self.failed.emit(traceback.format_exc())


class OcrWorker(QtCore.QObject):
    log = QtCore.Signal(str)
    done = QtCore.Signal(list)   # list[(idx, rec)]
    failed = QtCore.Signal(str)

    def __init__(self, items: list[EntryRecord], indices: list[int], download_dir: str):
        super().__init__()
        self.items = items
        self.indices = indices
        self.download_dir = download_dir
        self._stop = False

    def stop(self):
        self._stop = True

    def _read_image_bytes_for_record(self, sess: requests.Session, rec: EntryRecord) -> Optional[bytes]:
        if rec.bild_datei and os.path.isfile(rec.bild_datei):
            try:
                with open(rec.bild_datei, "rb") as f:
                    return f.read()
            except Exception:
                pass

        if not rec.bild_url:
            try:
                r = sess.get(rec.detail_url, timeout=25)
                if r.status_code == 200 and r.text:
                    soup = BeautifulSoup(r.text, "html.parser")
                    meta = soup.find("meta", attrs={"property": "og:image"})
                    if meta and meta.get("content"):
                        rec.bild_url = urljoin(rec.detail_url, meta["content"])
                    if not rec.bild_url:
                        meta = soup.find("meta", attrs={"name": "twitter:image"})
                        if meta and meta.get("content"):
                            rec.bild_url = urljoin(rec.detail_url, meta["content"])
                    if not rec.bild_url:
                        img = soup.find("img", src=True)
                        if img:
                            rec.bild_url = urljoin(rec.detail_url, img["src"])
            except Exception:
                pass

        if rec.bild_url:
            try:
                r = sess.get(rec.bild_url, timeout=20)
                if r.status_code == 200 and r.content:
                    if not rec.bild_datei:
                        folder = os.path.join(self.download_dir, "bilder")
                        os.makedirs(folder, exist_ok=True)
                        fn = f"{safe_filename(rec.nachname)}_{safe_filename(rec.name)}_{safe_filename(rec.geburtsdatum)}_{safe_filename(rec.sterbedatum)}.png"
                        fp = os.path.join(folder, fn)
                        try:
                            im = Image.open(BytesIO(r.content))
                            if im.mode in ("P",):
                                im = im.convert("RGBA")
                            if im.mode not in ("RGB", "RGBA"):
                                im = im.convert("RGBA")
                            im.save(fp, format="PNG", optimize=True)
                            rec.bild_datei = fp
                        except Exception:
                            pass
                    return r.content
            except Exception:
                return None
        return None

    @QtCore.Slot()
    def run(self):
        try:
            if not pytesseract:
                self.failed.emit("OCR nicht verfügbar: pytesseract ist nicht installiert.")
                return

            ok = configure_tesseract(self.log.emit)
            if not ok:
                self.failed.emit("OCR nicht verfügbar: tesseract.exe nicht gefunden/konfiguriert.")
                return

            sess = make_session()
            updated = []

            for idx in self.indices:
                if self._stop:
                    self.log.emit("OCR: Stop angefordert.")
                    break
                if idx < 0 or idx >= len(self.items):
                    continue

                rec = self.items[idx]
                display_name = rec.full_name or (rec.name + " " + rec.nachname).strip()
                self.log.emit(f"OCR: {display_name}")

                img_bytes = self._read_image_bytes_for_record(sess, rec)
                if not img_bytes:
                    self.log.emit("OCR: kein Bild gefunden – übersprungen.")
                    continue

                try:
                    text = ocr_best_effort(img_bytes)
                except Exception as e:
                    self.log.emit(f"OCR Fehler: {e}")
                    continue

                if not text:
                    self.log.emit("OCR: kein Text erkannt.")
                    continue

                if rec.zusatzinformationen:
                    if text not in rec.zusatzinformationen:
                        rec.zusatzinformationen = (rec.zusatzinformationen.rstrip() + "\n\n" + text).strip()
                else:
                    rec.zusatzinformationen = text

                updated.append((idx, rec))
                self.log.emit("OCR: Zusatzinfos aktualisiert.")

            self.done.emit(updated)

        except Exception:
            self.failed.emit(traceback.format_exc())


# ---------------- Helpers: Dialog ----------------

class TesseractHelpDialog(QtWidgets.QMessageBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Tesseract fehlt")
        self.setIcon(QtWidgets.QMessageBox.Warning)
        self.setTextFormat(QtCore.Qt.RichText)
        self.setText(
            "Tesseract OCR ist auf diesem System nicht gefunden worden.<br><br>"
            "Bitte installiere Tesseract (TU Mannheim Build) und starte das Programm danach neu.<br><br>"
            f'<a href="{TESSERACT_DOWNLOAD_URL}">{TESSERACT_DOWNLOAD_URL}</a>'
        )
        self.setStandardButtons(QtWidgets.QMessageBox.Ok)
        self.setTextInteractionFlags(QtCore.Qt.TextBrowserInteraction)
        self.setDefaultButton(QtWidgets.QMessageBox.Ok)
        self.setStyleSheet("QLabel { min-width: 560px; }")


# ---------------- UI ----------------

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Traueranzeigen Downloader (Multi-Sites)")
        self.resize(1420, 820)

        self.items: list[EntryRecord] = []

        self.thread: Optional[QtCore.QThread] = None
        self.worker: Optional[CrawlWorker] = None

        self.ocr_thread: Optional[QtCore.QThread] = None
        self.ocr_worker: Optional[OcrWorker] = None

        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        layout = QtWidgets.QVBoxLayout(central)

        # --- Controls
        grid = QtWidgets.QGridLayout()
        layout.addLayout(grid)

        self.url_edit = QtWidgets.QLineEdit()
        self.url_edit.setPlaceholderText(
            "trauer-anzeigen.de | abschied-nehmen.de | ok-trauer.de | gedenken.freiepresse.de | vrm-trauer.de | Max Mustermann"
        )
        grid.addWidget(QtWidgets.QLabel("URL oder Name:"), 0, 0)
        grid.addWidget(self.url_edit, 0, 1, 1, 7)

        self.dir_edit = QtWidgets.QLineEdit(os.path.abspath("downloads"))
        self.dir_btn = QtWidgets.QPushButton("Ordner …")
        self.dir_btn.clicked.connect(self.pick_dir)
        grid.addWidget(QtWidgets.QLabel("Zielordner:"), 1, 0)
        grid.addWidget(self.dir_edit, 1, 1, 1, 6)
        grid.addWidget(self.dir_btn, 1, 7)

        self.mode_combo = QtWidgets.QComboBox()
        self.mode_combo.addItem("Bilder", "images")
        self.mode_combo.addItem("Daten", "data")
        self.mode_combo.addItem("Bilder + Daten", "both")
        grid.addWidget(QtWidgets.QLabel("Download:"), 2, 0)
        grid.addWidget(self.mode_combo, 2, 1)

        self.max_people_spin = QtWidgets.QSpinBox()
        self.max_people_spin.setRange(0, 1_000_000)
        self.max_people_spin.setValue(0)
        self.max_people_spin.setToolTip("0 = unbegrenzt. Sobald erreicht, stoppt der Crawler automatisch.")
        grid.addWidget(QtWidgets.QLabel("Max. Personen:"), 2, 2)
        grid.addWidget(self.max_people_spin, 2, 3)

        self.export_combo = QtWidgets.QComboBox()
        self.export_combo.addItem("Excel (.xlsx)", "xlsx")
        self.export_combo.addItem("JSON (.json)", "json")
        self.export_combo.addItem("CSV (.csv)", "csv")
        self.export_combo.addItem("XML (.xml)", "xml")
        grid.addWidget(QtWidgets.QLabel("Daten speichern als:"), 2, 6)
        grid.addWidget(self.export_combo, 2, 7)

        # --- Buttons row + Search
        row = QtWidgets.QHBoxLayout()
        layout.addLayout(row)

        self.start_btn = QtWidgets.QPushButton("Start")
        self.stop_btn = QtWidgets.QPushButton("Stop")
        self.ocr_btn = QtWidgets.QPushButton("OCR für diese Person(en)")
        self.del_btn = QtWidgets.QPushButton("Auswahl löschen")
        self.reset_btn = QtWidgets.QPushButton("Reset")
        self.save_btn = QtWidgets.QPushButton("Daten speichern")

        self.stop_btn.setEnabled(False)

        self.start_btn.clicked.connect(self.start)
        self.stop_btn.clicked.connect(self.stop)
        self.save_btn.clicked.connect(self.save_data)
        self.ocr_btn.clicked.connect(self.run_ocr_for_selected)
        self.del_btn.clicked.connect(self.delete_selected_rows)
        self.reset_btn.clicked.connect(self.reset_all)

        row.addWidget(self.start_btn)
        row.addWidget(self.stop_btn)
        row.addWidget(self.ocr_btn)
        row.addWidget(self.del_btn)
        row.addWidget(self.reset_btn)
        row.addSpacing(16)

        row.addWidget(QtWidgets.QLabel("Suche Person:"))
        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("Filter in der Tabelle (Nachname, Ort, ...)")
        self.search_edit.textChanged.connect(self.apply_filter)
        row.addWidget(self.search_edit, 1)

        row.addSpacing(16)
        row.addWidget(self.save_btn)

        # --- Table
        self.table = QtWidgets.QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels([
            "Name", "Nachname", "PLZ", "Ort",
            "Geburtsdatum", "Sterbedatum",
            "Zusatzinformationen", "Anzeige-URL"
        ])
        self.table.setColumnWidth(6, 520)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.table.cellDoubleClicked.connect(self.open_row_url)

        # Sortierung per Klick auf Header
        self.table.setSortingEnabled(True)

        layout.addWidget(self.table, 2)

        # --- Log
        self.log = QtWidgets.QPlainTextEdit()
        self.log.setReadOnly(True)
        self.log.setMaximumBlockCount(5000)
        layout.addWidget(self.log, 1)

        # --- Shortcuts
        QShortcut(QKeySequence(QtCore.Qt.Key_Delete), self, activated=self.delete_selected_rows)
        QShortcut(QKeySequence("Ctrl+S"), self, activated=self.save_data)
        QShortcut(QKeySequence("Ctrl+Q"), self, activated=self.close)
        QShortcut(QKeySequence("Ctrl+A"), self, activated=self.select_all_rows)
        QShortcut(QKeySequence("Ctrl+O"), self, activated=self.run_ocr_for_selected)

        # --- Style
        self.setStyleSheet("""
            QMainWindow { background: #0f1115; }
            QLabel, QLineEdit, QComboBox, QPlainTextEdit, QTableWidget, QSpinBox { color: #e7e7e7; }
            QLineEdit, QComboBox, QPlainTextEdit, QTableWidget, QSpinBox {
                background: #151922; border: 1px solid #262b36; border-radius: 10px; padding: 6px;
            }
            QPushButton {
                background: #2a3242; color: #fff; border: 1px solid #3a445a;
                border-radius: 12px; padding: 8px 14px;
            }
            QPushButton:hover { background: #36405a; }
            QPushButton:disabled { background: #202634; color: #888; border-color: #202634; }
            QHeaderView::section { background: #151922; color: #e7e7e7; border: 1px solid #262b36; padding: 6px; }
        """)

    # ---------- UI helpers ----------

    def append_log(self, msg: str):
        self.log.appendPlainText(msg)

    def pick_dir(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "Zielordner wählen", self.dir_edit.text())
        if d:
            self.dir_edit.setText(d)

    # ---------- Selection helpers ----------

    def _selected_item_indices(self) -> List[int]:
        rows = sorted({i.row() for i in self.table.selectionModel().selectedRows()})
        return [r for r in rows if 0 <= r < len(self.items)]

    def select_all_rows(self):
        if self.table.rowCount() == 0:
            return
        self.table.selectAll()

    # ---------- Delete / Reset ----------

    def delete_selected_rows(self):
        indices = self._selected_item_indices()
        if not indices:
            return
        for idx in sorted(indices, reverse=True):
            if 0 <= idx < len(self.items):
                del self.items[idx]
        self.render_table()
        self.apply_filter(self.search_edit.text())
        self.append_log(f"Gelöscht: {len(indices)} Eintrag/Einträge")

    def reset_all(self):
        self.stop()
        self.url_edit.clear()
        self.search_edit.clear()
        self.max_people_spin.setValue(0)
        self.items = []
        self.table.setRowCount(0)
        self.log.clear()
        self.append_log("Reset: Ausgangszustand wiederhergestellt.")

    # ---------- Crawl ----------

    def start(self):
        txt = (self.url_edit.text() or "").strip()
        if not txt:
            QtWidgets.QMessageBox.information(self, "Fehlt", "Bitte URL oder Name eingeben.")
            return

        mode = self.mode_combo.currentData()
        out_dir = self.dir_edit.text().strip() or os.path.abspath("downloads")
        max_people = int(self.max_people_spin.value())

        self.items = []
        self.table.setRowCount(0)

        if re.match(r"^https?://", txt, flags=re.I):
            self.append_log(f"Starte URL-Crawl… (Modus: {mode})")
        else:
            self.append_log(f"Starte Name-Suche über alle Seiten… (Name: {txt}, Modus: {mode})")

        opts = CrawlOptions(mode=mode, download_dir=out_dir, max_people=max_people)

        self.thread = QtCore.QThread()
        self.worker = CrawlWorker(txt, opts)
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.worker.log.connect(self.append_log)
        self.worker.done.connect(self.on_done)
        self.worker.failed.connect(self.on_failed)

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.thread.start()

    def stop(self):
        if self.worker:
            self.worker.stop()
            self.append_log("Stop angefordert…")
        if self.ocr_worker:
            self.ocr_worker.stop()
            self.append_log("OCR Stop angefordert…")

    def on_done(self, items: list):
        self.items = items
        self.append_log(f"Fertig. Einträge: {len(items)}")
        self.render_table()

        self.stop_btn.setEnabled(False)
        self.start_btn.setEnabled(True)

        if self.thread:
            self.thread.quit()
            self.thread.wait()
        self.thread = None
        self.worker = None

        self.apply_filter(self.search_edit.text())

    def on_failed(self, tb_str: str):
        self.append_log(tb_str)
        QtWidgets.QMessageBox.critical(self, "Fehler", "Crawler ist abgestürzt – siehe Log.")
        self.stop_btn.setEnabled(False)
        self.start_btn.setEnabled(True)

        if self.thread:
            self.thread.quit()
            self.thread.wait()
        self.thread = None
        self.worker = None

    # ---------- Table + Search ----------

    def render_table(self):
        self.table.setSortingEnabled(False)  # während wir füllen
        self.table.setRowCount(len(self.items))
        for row, r in enumerate(self.items):
            self._render_row(row, r)
        self.table.setSortingEnabled(True)
        self.table.resizeColumnsToContents()

    def _render_row(self, row: int, r: EntryRecord):
        def setc(col: int, val: str, sort_key: Optional[str] = None):
            it = QtWidgets.QTableWidgetItem(val or "")
            # WICHTIG: Sortierung nutzt EditRole (damit Qt wirklich danach sortiert)
            if sort_key is not None:
                it.setData(QtCore.Qt.EditRole, sort_key)
            self.table.setItem(row, col, it)

        def date_key(d: str) -> str:
            d = normalize_date_de(d)
            m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", d)
            if not m:
                return ""
            dd, mm, yyyy = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
            return f"{yyyy}-{mm}-{dd}"

        setc(0, r.name, (r.name or "").lower())
        setc(1, r.nachname, (r.nachname or "").lower())
        setc(2, r.plz, re.sub(r"\D+", "", r.plz or ""))
        setc(3, r.ort, (r.ort or "").lower())
        setc(4, r.geburtsdatum, date_key(r.geburtsdatum))
        setc(5, r.sterbedatum, date_key(r.sterbedatum))
        setc(6, r.zusatzinformationen, (r.zusatzinformationen or "").lower())
        setc(7, r.detail_url, r.detail_url or "")

    def apply_filter(self, text: str):
        q = (text or "").strip().lower()
        for row in range(self.table.rowCount()):
            if not q:
                self.table.setRowHidden(row, False)
                continue
            joined = " ".join(
                self.table.item(row, col).text() if self.table.item(row, col) else ""
                for col in range(self.table.columnCount())
            ).lower()
            self.table.setRowHidden(row, q not in joined)

    def open_row_url(self, row: int, col: int):
        it = self.table.item(row, 7)
        if it and it.text().strip():
            QDesktopServices.openUrl(QUrl(it.text().strip()))

    # ---------- Manual OCR ----------

    def _has_tesseract_available(self) -> bool:
        if not pytesseract:
            return False
        if shutil.which("tesseract"):
            return True
        for c in TESSERACT_CANDIDATES:
            if os.path.isfile(c):
                return True
        return False

    def _show_tesseract_help(self):
        dlg = TesseractHelpDialog(self)
        dlg.exec()

    def run_ocr_for_selected(self):
        indices = self._selected_item_indices()
        if not indices:
            QtWidgets.QMessageBox.information(self, "Keine Auswahl", "Bitte eine oder mehrere Personen (Zeilen) auswählen.")
            return

        if not self._has_tesseract_available():
            self._show_tesseract_help()
            return

        out_dir = self.dir_edit.text().strip() or os.path.abspath("downloads")

        if self.ocr_worker:
            self.ocr_worker.stop()

        self.append_log(f"OCR startet für {len(indices)} Person(en)…")

        self.ocr_thread = QtCore.QThread()
        self.ocr_worker = OcrWorker(self.items, indices, out_dir)
        self.ocr_worker.moveToThread(self.ocr_thread)

        self.ocr_thread.started.connect(self.ocr_worker.run)
        self.ocr_worker.log.connect(self.append_log)
        self.ocr_worker.done.connect(self.on_ocr_done)
        self.ocr_worker.failed.connect(self.on_ocr_failed)

        self.ocr_btn.setEnabled(False)
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)

        self.ocr_thread.start()

    def on_ocr_done(self, updated: list):
        for idx, rec in updated:
            if 0 <= idx < len(self.items):
                self.items[idx] = rec
                self._render_row(idx, rec)

        self.append_log(f"OCR fertig. Aktualisiert: {len(updated)}")
        self.table.resizeColumnsToContents()
        self.apply_filter(self.search_edit.text())

        self.ocr_btn.setEnabled(True)
        self.start_btn.setEnabled(True)

        if self.ocr_thread:
            self.ocr_thread.quit()
            self.ocr_thread.wait()
        self.ocr_thread = None
        self.ocr_worker = None

    def on_ocr_failed(self, msg: str):
        self.append_log(msg)
        QtWidgets.QMessageBox.critical(self, "OCR Fehler", msg)

        self.ocr_btn.setEnabled(True)
        self.start_btn.setEnabled(True)

        if self.ocr_thread:
            self.ocr_thread.quit()
            self.ocr_thread.wait()
        self.ocr_thread = None
        self.ocr_worker = None

    # ---------- Save / Export ----------

    def save_data(self):
        if not self.items:
            QtWidgets.QMessageBox.information(self, "Keine Daten", "Noch keine Einträge.")
            return

        fmt = self.export_combo.currentData()
        filters = {
            "xlsx": "Excel (*.xlsx)",
            "json": "JSON (*.json)",
            "csv": "CSV (*.csv)",
            "xml": "XML (*.xml)",
        }[fmt]

        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Speichern", "", filters)
        if not path:
            return

        try:
            if fmt == "xlsx":
                export_xlsx(path, self.items)
            elif fmt == "json":
                export_json(path, self.items)
            elif fmt == "csv":
                export_csv(path, self.items)
            elif fmt == "xml":
                export_xml(path, self.items)
            self.append_log(f"Daten gespeichert: {path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Fehler", str(e))


def main():
    app = QtWidgets.QApplication([])
    w = MainWindow()
    w.show()
    app.exec()


if __name__ == "__main__":
    main()
